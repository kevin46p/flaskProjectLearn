import datetime
import os
from typing import List, Dict, Any
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment

from app.configs.constant import DIR_PATH, UPLOAD_PATH
from app.utils.helpers import get_uuid
from collections import namedtuple

Title = namedtuple("Title", ["start", "end", "name", "field_name"])


class OrderExportMeta(type):
    def __init__(cls, what, base, dict):
        cells = []

        fields = {}

        for k, v in dict.items():
            if isinstance(v, Title):
                cells.append(v)

                if v.field_name is not None:
                    fields[v.field_name] = v

        cls.__cells__ = cells
        cls.__fields__ = fields

        super().__init__(what, base, dict)


class OrderExport(metaclass=OrderExportMeta):
    file_dir = "default"

    def __init__(self):

        self.wb = Workbook()
        self.ws = self.wb.active
        self.th = datetime.datetime.now()

        self.file_name = f"{get_uuid()}.xlsx"
        self.relative_path = f"/export/{self.file_dir}/{self.th.year}/{str(self.th.month).zfill(2)}"
        self.path = os.path.abspath(DIR_PATH + UPLOAD_PATH + self.relative_path)

        if not os.path.exists(self.path):
            os.makedirs(self.path)

    @property
    def fields(self):
        return self.__fields__

    @property
    def cells(self):
        return self.__cells__

    def create_excel_title(self):

        for cell in self.cells:
            if cell.start != cell.end:
                self.ws.merge_cells(f"{cell.start}:{cell.end}")
            th_cell = self.ws[cell.start]
            th_cell.value = cell.name
            th_cell.alignment = Alignment(horizontal='center', vertical='center')

    def save(self, datas: List[Dict[str, Any]], start_row=2, sub_path="inventory_order"):
        self.create_excel_title()

        for data in datas:
            for k, v in data.items():
                title = self.fields.get(k)

                if title:
                    self.ws[f"{title.start[:-1]}{str(start_row)}"] = v
            start_row += 1

        path = os.path.abspath(self.path + "/" + self.file_name)
        self.wb.save(path)
        self.wb.close()

        return f"{current_app.config['EXPORT_FILE_HOST']}/dms-app-assets-ms{self.relative_path}/{self.file_name}"


class InventoryExport(OrderExport):
    file_dir = 'inventory_order/inventory_list'

    A1_A2 = Title(start="A1", end="A2", name="盘点日期", field_name="createTime")
    B1_B2 = Title(start="B1", end="B2", name="盘点单状态", field_name="orderStatusName")
    C1_C2 = Title(start="C1", end="C2", name="盘点单号", field_name="inventoryOrderNo")
    D1_D2 = Title(start="D1", end="D2", name="会签进度", field_name="signStatusName")
    E1_E2 = Title(start="E1", end="E2", name="会签完成时间", field_name="signFinishTime")
    F1_F2 = Title(start="F1", end="F2", name="制单人", field_name="createUserName")
    G1_I1 = Title(start="G1", end="I1", name="应盘", field_name=None)
    G2_G2 = Title(start="G2", end="G2", name="数量", field_name="inventorySum")

    H2_H2 = Title(start="H2", end="H2", name="含税金额", field_name="inventoryTotalAmount")
    I2_I2 = Title(start="I2", end="I2", name="不含税金额", field_name="inventoryTotalAmountNoTax")
    J1_L1 = Title(start="J1", end="L1", name="正常", field_name=None)
    J2_J2 = Title(start="J2", end="J2", name="数量", field_name="idNormalNum")
    K2_K2 = Title(start="K2", end="K2", name="含税金额", field_name="idNormalTotalAmount")
    L2_L2 = Title(start="L2", end="L2", name="不含税金额", field_name="idNormalTotalAmountNoTax")
    M1_O1 = Title(start="M1", end="O1", name="部分损坏", field_name=None)
    M2_M2 = Title(start="M2", end="M2", name="数量", field_name="idPartDamageNum")
    N2_N2 = Title(start="N2", end="N2", name="含税金额", field_name="idPartDamageTotalAmount")
    O2_O2 = Title(start="O2", end="O2", name="不含税金额", field_name="idPartDamageTotalAmountNoTax")
    P1_R1 = Title(start="P1", end="R1", name="不可使用", field_name=None)
    P2_P2 = Title(start="P2", end="P2", name="数量", field_name="idDisableNum")
    Q2_Q2 = Title(start="Q2", end="Q2", name="含税金额", field_name="idDisableTotalAmount")
    R2_R2 = Title(start="R2", end="R2", name="不含税金额", field_name="idDisableTotalAmountNoTax")
    S1_U1 = Title(start="S1", end="U1", name="丢失", field_name=None)
    S2_S2 = Title(start="S2", end="S2", name="数量", field_name="idLoseNum")
    T2_T2 = Title(start="T2", end="T2", name="含税金额", field_name="idLoseTotalAmount")
    U2_U2 = Title(start="U2", end="U2", name="不含税金额", field_name="idLoseTotalAmountNoTax")
    V1_W1 = Title(start="V1", end="W1", name="岗位盘点数量", field_name=None)
    V2_V2 = Title(start="V2", end="V2", name="财务岗", field_name="financialInventoryNum")
    W2_W2 = Title(start="W2", end="W2", name="业务岗", field_name="businessInventoryNum")


class InventoryDetailListExport(OrderExport):
    file_dir = 'inventory_order/inventory_detail_list'

    A1_A3 = Title(start="A1", end="A1", name="资产名称", field_name="atomName")
    B1_B3 = Title(start="B1", end="B1", name="资产编码", field_name="atomCode")
    C1_C3 = Title(start="C1", end="C1", name="资产类别三级", field_name="thirdCategoryName")

    D1_D3 = Title(start="D1", end="D1", name="品牌", field_name="brandName")
    E1_E3 = Title(start="E1", end="E1", name="规格型号", field_name="skuName")
    F1_F3 = Title(start="F1", end="F1", name="单位", field_name="unitName")
    G1_G3 = Title(start="G1", end="G1", name="原使用状态", field_name="useStatusName")

    H1_H3 = Title(start="H1", end="H1", name="使用性质", field_name="usePropertyName")
    I1_I3 = Title(start="I1", end="I1", name="使用部门", field_name="useDepartmentName")
    J1_J3 = Title(start="J1", end="J1", name="使用人", field_name="recipientPersonName")
    K1_K3 = Title(start="K1", end="K1", name="责任人", field_name="responsiblePersonName")
    L1_L3 = Title(start="L1", end="L1", name="应盘数量", field_name="inventoryNum")
    M1_M3 = Title(start="M1", end="M1", name="应盘金额", field_name="inventoryAmount")
    N1_N3 = Title(start="N1", end="N1", name="盘点库存状态", field_name="inventoryStockStatusName")
    O1_O3 = Title(start="O1", end="O1", name="盘点使用状态", field_name="inventoryUseStatusName")
    P1_P3 = Title(start="P1", end="P1", name="操作", field_name=None)
    Q1_Q3 = Title(start="Q1", end="Q1", name="备注", field_name="remark")


class PurchaseOrderListExport(OrderExport):
    file_dir = 'purchase_order/purchase_order_list'
    A1_A1 = Title(start="A1", end="A1", name="采购单号", field_name="purchaseNo")
    B1_B1 = Title(start="B1", end="B1", name="制单日期", field_name="purchaseDate")
    C1_C1 = Title(start="C1", end="C1", name="单据状态", field_name="orderStatus")

    D1_D1 = Title(start="D1", end="D1", name="店铺编号", field_name="companyCode")
    E1_E1 = Title(start="E1", end="E1", name="店铺名称", field_name="companyName")
    F1_F1 = Title(start="F1", end="F1", name="店铺品牌", field_name="companyBrand")
    G1_G1 = Title(start="G1", end="G1", name="事业部", field_name="businessUnit")

    H1_H1 = Title(start="H1", end="H1", name="购买渠道", field_name="purchaseChannel")
    I1_I1 = Title(start="I1", end="I1", name="购买类别", field_name="purchaseClass")
    J1_J1 = Title(start="J1", end="J1", name="供应商", field_name="supplierName")

    K1_K1 = Title(start="K1", end="K1", name="供应商编码", field_name="supplierCode")
    L1_L1 = Title(start="L1", end="L1", name="供应商纳税人识别号", field_name="supplierTaxpayerID")
    N1_N1 = Title(start="M1", end="M1", name="是否补录", field_name="isReplenish")
    O1_O1 = Title(start="N1", end="N1", name="申请人", field_name="applicantPerson")
    P1_P1 = Title(start="O1", end="O1", name="采购申请数量", field_name="computePurchaseApplyNum")
    Q1_Q1 = Title(start="P1", end="P1", name="采购金额(含税)", field_name="computePurchaseApplyAmountWithTax")
    R1_R1 = Title(start="Q1", end="Q1", name="已入库数量", field_name="computeStockInNum")
    S1_S1 = Title(start="R1", end="R1", name="已入库金额", field_name="computeStockInAmount")
    T1_T1 = Title(start="S1", end="S1", name="未入库数量", field_name="computeUnStockInNum")
    U1_U1 = Title(start="T1", end="T1", name="未入库金额", field_name="computeUnStockInAmount")
    X1_X1 = Title(start="U1", end="U1", name="已付款金额", field_name="computePaidAmount")
    Y1_Y1 = Title(start="V1", end="V1", name="未付款金额", field_name="computeUnPaidAmount")


class StockListExport(OrderExport):
    file_dir = 'stock/stock_list'
    A1_A1 = Title(start="A1", end="A1", name="资产编码", field_name="assetsCode")
    B1_B1 = Title(start="B1", end="B1", name="权属公司编号", field_name="belongCompanyCode")
    C1_C1 = Title(start="C1", end="C1", name="权属公司简称", field_name="belongCompanyName")
    D1_D1 = Title(start="D1", end="D1", name="使用公司编码", field_name="useCompanyCode")
    E1_E1 = Title(start="E1", end="E1", name="使用公司简称", field_name="useCompanyName")

    F1_F1 = Title(start="F1", end="F1", name="资产类别三级编码", field_name="thirdCategoryCode")
    G1_G1 = Title(start="G1", end="G1", name="资产类别三级", field_name="thirdCategoryName")
    H1_H1 = Title(start="H1", end="H1", name="资产名称", field_name="assetsName")
    I1_I1 = Title(start="I1", end="I1", name="供应商", field_name="supplierName")
    J1_J1 = Title(start="J1", end="J1", name="品牌", field_name="brandName")
    K1_K1 = Title(start="K1", end="K1", name="规格型号", field_name="assetsSKUName")
    L1_L1 = Title(start="L1", end="L1", name="原厂编码（序列号）", field_name="originalFactoryCode")
    M1_M1 = Title(start="M1", end="M1", name="单位", field_name="unitName")
    N1_N1 = Title(start="N1", end="N1", name="库存数量", field_name="assetsNum")
    O1_O1 = Title(start="O1", end="O1", name="发票金额（不含税）", field_name="invoiceAmount")
    P1_P1 = Title(start="P1", end="P1", name="安装费（不含税）", field_name="installAmount")
    Q1_Q1 = Title(start="Q1", end="Q1", name="入账金额（不含税）", field_name="inAccountAmount")
    R1_R1 = Title(start="R1", end="R1", name="入账金额（含税）", field_name="inAccountTaxAmount")
    S1_S1 = Title(start="S1", end="S1", name="开始使用时间", field_name="startUseTime")
    T1_T1 = Title(start="T1", end="T1", name="购入时间", field_name="purchaseTime")
    U1_U1 = Title(start="U1", end="U1", name="入库时间", field_name="stockInTime")
    V1_V1 = Title(start="V1", end="V1", name="处置/出库时间", field_name="stockOutTime")
    W1_W1 = Title(start="W1", end="W1", name="创建时间", field_name="dataCreateTime")
    X1_X1 = Title(start="X1", end="X1", name="库存状态", field_name="stockStatusName")
    Y1_Y1 = Title(start="Y1", end="Y1", name="使用状态", field_name="useStatusName")
    Z1_Z1 = Title(start="Z1", end="Z1", name="使用性质", field_name="usePropertyName")
    AA1_AA1 = Title(start="AA1", end="AA1", name="使用部门", field_name="useDepartmentName")
    AB1_AB1 = Title(start="AB1", end="AB1", name="使用人", field_name="recipientPerson")
    AC1_AC1 = Title(start="AC1", end="AC1", name="责任人", field_name="responsiblePerson")
    AD1_AD1 = Title(start="AD1", end="AD1", name="存放地点", field_name="storagePlace")
    AE1_AE1 = Title(start="AE1", end="AE1", name="备注", field_name="remark")
    AF1_AF1 = Title(start="AF1", end="AF1", name="入库类型", field_name="stockInTypeName")
    AG1_AG1 = Title(start="AG1", end="AG1", name="图片数量", field_name="stockPicNum")


class SupplierListExport(OrderExport):
    """供应商列表导出"""
    file_dir = 'supplier/supplier_list'
    A1_A1 = Title(start="A1", end="A1", name="供应商名称", field_name="name")
    B1_B1 = Title(start="B1", end="B1", name="供应商代码", field_name="code")
    C1_C1 = Title(start="C1", end="C1", name="纳税人识别号", field_name="taxpayerId")
    D1_D1 = Title(start="D1", end="D1", name="新增资产时是否可选", field_name="isCanUseText")
    E1_E1 = Title(start="E1", end="E1", name="新增时间", field_name="createTime")


class ReportDamageListExport(OrderExport):
    """报损单列表导出"""
    file_dir = 'report_damage/report_damage_list'
    A1_A3 = Title(start="A1", end="A3", name="店铺简称", field_name="companyName")
    B1_B3 = Title(start="B1", end="B3", name="盘点单号", field_name="inventoryOrderNo")
    C1_G1 = Title(start="C1", end="G1", name="报损", field_name=None)
    C2_C3 = Title(start="C2", end="C3", name="报损单号", field_name="reportDamageCode")
    D2_E2 = Title(start="D2", end="E2", name="报废", field_name=None)
    D3_D3 = Title(start="D3", end="D3", name="数量", field_name="allDisposeScrapNum")
    E3_E3 = Title(start="E3", end="E3", name="金额(含税)", field_name="allDisposeScrapAmount")
    F2_G2 = Title(start="F2", end="G2", name="丢失", field_name=None)
    F3_F3 = Title(start="F3", end="F3", name="数量", field_name="allLoseNum")
    G3_G3 = Title(start="G3", end="G3", name="金额(含税)", field_name="allLoseAmount")
    H1_N1 = Title(start="H1", end="N1", name="审批", field_name=None)
    H2_H3 = Title(start="H2", end="H3", name="审批单号", field_name="approveOrderNo")
    I2_J2 = Title(start="I2", end="J2", name="报废", field_name=None)
    I3_I3 = Title(start="I3", end="I3", name="数量", field_name="approveDisposeScrapNum")
    J3_J3 = Title(start="J3", end="J3", name="金额(含税)", field_name="approveDisposeScrapAmount")
    K2_L2 = Title(start="K2", end="L2", name="丢失", field_name=None)
    K3_K3 = Title(start="K3", end="K3", name="数量", field_name="approveLoseNum")
    L3_L3 = Title(start="L3", end="L3", name="金额(含税)", field_name="approveLoseAmount")
    M2_M3 = Title(start="M2", end="M3", name="处置金额", field_name="approveDisposeAmount")
    N2_N3 = Title(start="N2", end="N3", name="赔偿金额", field_name="approveCompensationAmount")

    O1_O3 = Title(start="O1", end="O3", name="报损单状态", field_name="orderStatusName")

    P1_P3 = Title(start="P1", end="P3", name="审批单状态", field_name="approveStatusName")

    Q1_Q3 = Title(start="Q1", end="Q3", name="审批单终结日期", field_name="approveFinishTime")
    R1_R3 = Title(start="R1", end="R3", name="申请人", field_name="applicantPerson")
    S1_S3 = Title(start="S1", end="S3", name="店铺品牌", field_name="compBrandName")
    T1_T3 = Title(start="T1", end="T3", name="事业部", field_name="businessName")
    U1_U3 = Title(start="U1", end="U3", name="店铺编码", field_name="companyCode")


class RepeatPerson(OrderExport):
    """供应商列表导出"""
    file_dir = 'person/person_list'
    A1_A1 = Title(start="A1", end="A1", name="资产编码", field_name="atom_code")
    B1_B1 = Title(start="B1", end="B1", name="领用人uid", field_name="recipientUid")
    C1_C1 = Title(start="C1", end="C1", name="领用人psnCode", field_name="recipientPsnCode")
    D1_D1 = Title(start="D1", end="D1", name="领用人名字", field_name="recipientPerson")
    E1_E1 = Title(start="E1", end="E1", name="责任人uid", field_name="responsibleUid")
    F1_F1 = Title(start="F1", end="F1", name="责任人psnCode", field_name="responsiblePsnCode")
    G1_G1 = Title(start="G1", end="G1", name="责任人名字", field_name="responsiblePerson")
